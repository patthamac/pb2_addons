# -*- coding: utf-8 -*-
# © <YEAR(S)> <AUTHOR(S)>
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl.html).
import base64
import openpyxl
import cStringIO
from openerp import models, fields, api, _
from openerp.exceptions import Warning as UserError
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles.colors import Color


class BudgetExportWizard(models.Model):
    _name = 'unit.budget.plan.export'

    attachment_id = fields.Many2one('ir.attachment', 'Template')
    editable_lines = fields.Integer('Editable lines', required=True, default=10)

    @api.multi
    def update_budget_xls(self, budget_ids, template_id=None):
        if not template_id:
            raise UserError(_('Please add .xlsx template.'))
        if not budget_ids:
            raise UserError(_('No budget to export.'))
        budgets = self.env['budget.plan.unit'].browse(budget_ids)
        template_file = self.env['ir.attachment'].browse(template_id)

        for budget in budgets:
            exp_file = self.attachment_id.datas.decode('base64')
            stream = cStringIO.StringIO(exp_file)
            workbook = openpyxl.load_workbook(stream)
            worksheets = workbook.worksheets
            
            Activity_Group_Sheet = workbook.get_sheet_by_name('ActivityGroup_MasterData')
            Activity_Group_Sheet.protection.sheet = True
            activities = self.env['account.activity.group'].search([])

            Activity_Group_Sheet.cell(row=1, column=1, value='Sequence')
            Activity_Group_Sheet.cell(row=1, column=2, value='Activity Group - English')
            Activity_Group_Sheet.cell(row=1, column=3, value='Activity Group - Thai')

            ag_row = 2
            ag_count = 1
            for ag in activities:
                Activity_Group_Sheet.cell(row=ag_row, column=1, value=ag_count)
                Activity_Group_Sheet.cell(row=ag_row, column=2, value=ag.name)
                Activity_Group_Sheet.cell(row=ag_row, column=3, value=ag.name)
                ag_row += 1
                ag_count += 1
            
            formula1 = "{0}!$C$2:$C$%s" %(ag_row)
            
            dv = DataValidation(type="list",
                    formula1=formula1.format(quote_sheetname('ActivityGroup_MasterData'))
                    )
            
            Non_CostControl_Sheet = workbook.get_sheet_by_name('Non_CostControl')
            Non_CostControl_Sheet.protection.sheet = True
            
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            redFill = PatternFill(start_color='D3D3D3',
                   end_color='D3D3D3',
                   fill_type='solid')

            Non_CostControl_Sheet.add_data_validation(dv)
            Non_CostControl_Sheet.cell(row=1, column=2, value=budget.fiscalyear_id.name)
            Non_CostControl_Sheet.cell(row=2, column=2, value=budget.org_id.code)
            Non_CostControl_Sheet.cell(row=3, column=2, value=budget.section_id.code)
            Non_CostControl_Sheet.cell(row=4, column=2, value=fields.Date.today())
            Non_CostControl_Sheet.cell(row=5, column=2, value=self.env.user.name)
            Non_CostControl_Sheet.cell(row=6, column=2, value=str(budget.planned_overall))
            
            row = 11
            section_name = ''
            protection = Protection(locked=False)
            line_start = row
            for line in budget.plan_line_ids:
                section_name = line.section_id.name
                
                Non_CostControl_Sheet.cell(row=row, column=1, value="=$B$2").fill = redFill
                Non_CostControl_Sheet.cell(row=row, column=1).border = border
                Non_CostControl_Sheet.cell(row=row, column=2, value="=$B$3").fill = redFill
                Non_CostControl_Sheet.cell(row=row, column=2).border = border
                Non_CostControl_Sheet.cell(row=row, column=3, value=line.section_id.name).border = border
                Non_CostControl_Sheet.cell(row=row, column=3).fill = redFill
                Non_CostControl_Sheet.cell(row=row, column=4).protection =protection
                dv.add(Non_CostControl_Sheet.cell(row=row, column=4))
                ag_name = line.activity_group_id.name
                Non_CostControl_Sheet.cell(row=row, column=4, value=ag_name).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=4).border = border
                Non_CostControl_Sheet.cell(row=row, column=5).border = border
                Non_CostControl_Sheet.cell(row=row, column=6).border = border
                Non_CostControl_Sheet.cell(row=row, column=7).border = border
                Non_CostControl_Sheet.cell(row=row, column=8).border = border
                Non_CostControl_Sheet.cell(row=row, column=9).border = border

                Non_CostControl_Sheet.cell(row=row, column=7).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=8).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=9).protection = protection

                Non_CostControl_Sheet.cell(row=row, column=10, value="=G%s*H%s*I%s" %(row, row, row)).border = border
                Non_CostControl_Sheet.cell(row=row, column=10).fill = redFill

                Non_CostControl_Sheet.cell(row=row, column=11, value=line.m0).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=11).border = border
                Non_CostControl_Sheet.cell(row=row, column=11).fill = redFill
                
                Non_CostControl_Sheet.cell(row=row, column=12, value=line.m1).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=12).border = border
                
                Non_CostControl_Sheet.cell(row=row, column=13, value=line.m2).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=13).border = border
                
                Non_CostControl_Sheet.cell(row=row, column=14, value=line.m3).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=14).border = border

                Non_CostControl_Sheet.cell(row=row, column=15, value=line.m4).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=15).border = border

                Non_CostControl_Sheet.cell(row=row, column=16, value=line.m5).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=16).border = border

                Non_CostControl_Sheet.cell(row=row, column=17, value=line.m6).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=17).border = border

                Non_CostControl_Sheet.cell(row=row, column=18, value=line.m7).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=18).border = border
                
                Non_CostControl_Sheet.cell(row=row, column=19, value=line.m8).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=19).border = border
                
                Non_CostControl_Sheet.cell(row=row, column=20, value=line.m9).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=20).border = border
                
                Non_CostControl_Sheet.cell(row=row, column=21, value=line.m10).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=21).border = border
                
                Non_CostControl_Sheet.cell(row=row, column=22, value=line.m11).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=22).border = border
                
                Non_CostControl_Sheet.cell(row=row, column=23, value=line.m12).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=23).border = border
                
                Non_CostControl_Sheet.cell(row=row, column=24, value="=SUM(L%s:W%s)" %(row, row))
                Non_CostControl_Sheet.cell(row=row, column=24).border = border
                Non_CostControl_Sheet.cell(row=row, column=24).fill = redFill

                Non_CostControl_Sheet.cell(row=row, column=25, value="=J%s-X%s" %(row, row))
                Non_CostControl_Sheet.cell(row=row, column=25).border = border
                Non_CostControl_Sheet.cell(row=row, column=25).fill = redFill

                row += 1
                
            to_row = row + self.editable_lines
            
            for r in range(row, to_row):
                Non_CostControl_Sheet.cell(row=row, column=1, value="=$B$2").fill = redFill
                Non_CostControl_Sheet.cell(row=row, column=1).border = border
                Non_CostControl_Sheet.cell(row=row, column=2, value="=$B$3").fill = redFill
                Non_CostControl_Sheet.cell(row=row, column=2).border = border
                Non_CostControl_Sheet.cell(row=row, column=3, value=section_name).fill = redFill
                Non_CostControl_Sheet.cell(row=row, column=3).border = border
                
                dv.add(Non_CostControl_Sheet.cell(row=row, column=4))
                Non_CostControl_Sheet.cell(row=row, column=4).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=4).border = border
                Non_CostControl_Sheet.cell(row=row, column=5).border = border
                Non_CostControl_Sheet.cell(row=row, column=6).border = border
                Non_CostControl_Sheet.cell(row=row, column=7).border = border
                Non_CostControl_Sheet.cell(row=row, column=8).border = border
                Non_CostControl_Sheet.cell(row=row, column=9).border = border
                
                Non_CostControl_Sheet.cell(row=row, column=7).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=8).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=9).protection = protection
                
                Non_CostControl_Sheet.cell(row=row, column=10, value="=G%s*H%s*I%s" %(row, row, row)).border = border
                Non_CostControl_Sheet.cell(row=row, column=10).fill = redFill
                
                Non_CostControl_Sheet.cell(row=row, column=11).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=11).fill = redFill
                Non_CostControl_Sheet.cell(row=row, column=12).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=13).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=14).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=15).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=16).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=17).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=18).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=19).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=20).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=21).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=22).protection = protection
                Non_CostControl_Sheet.cell(row=row, column=23).protection = protection
                
                Non_CostControl_Sheet.cell(row=row, column=11).border = border
                Non_CostControl_Sheet.cell(row=row, column=12).border = border
                Non_CostControl_Sheet.cell(row=row, column=13).border = border
                Non_CostControl_Sheet.cell(row=row, column=14).border = border
                Non_CostControl_Sheet.cell(row=row, column=15).border = border
                Non_CostControl_Sheet.cell(row=row, column=16).border = border
                Non_CostControl_Sheet.cell(row=row, column=17).border = border
                Non_CostControl_Sheet.cell(row=row, column=18).border = border
                Non_CostControl_Sheet.cell(row=row, column=19).border = border
                Non_CostControl_Sheet.cell(row=row, column=20).border = border
                Non_CostControl_Sheet.cell(row=row, column=21).border = border
                Non_CostControl_Sheet.cell(row=row, column=22).border = border
                Non_CostControl_Sheet.cell(row=row, column=23).border = border

                Non_CostControl_Sheet.cell(row=row, column=24, value="=SUM(L%s:W%s)" %(row, row))
                Non_CostControl_Sheet.cell(row=row, column=24).border = border
                Non_CostControl_Sheet.cell(row=row, column=24).fill = redFill

                Non_CostControl_Sheet.cell(row=row, column=25, value="=J%s-X%s" %(row, row))
                Non_CostControl_Sheet.cell(row=row, column=25).border = border
                Non_CostControl_Sheet.cell(row=row, column=25).fill = redFill

                row += 1
                
            Non_CostControl_Sheet.cell(row=row, column=9).value = 'Total'
            Non_CostControl_Sheet.cell(row=row, column=9).border = border
            Non_CostControl_Sheet.cell(row=row, column=9).fill = redFill
            
            Non_CostControl_Sheet.cell(row=row, column=10).border = border
            Non_CostControl_Sheet.cell(row=row, column=10).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=10).value = '=SUM(J%s:J%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=11).border = border
            Non_CostControl_Sheet.cell(row=row, column=11).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=11).value = '=SUM(K%s:K%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=12).border = border
            Non_CostControl_Sheet.cell(row=row, column=12).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=12).value = '=SUM(L%s:L%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=13).border = border
            Non_CostControl_Sheet.cell(row=row, column=13).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=13).value = '=SUM(M%s:M%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=14).border = border
            Non_CostControl_Sheet.cell(row=row, column=14).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=14).value = '=SUM(N%s:N%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=15).border = border
            Non_CostControl_Sheet.cell(row=row, column=15).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=15).value = '=SUM(O%s:O%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=16).border = border
            Non_CostControl_Sheet.cell(row=row, column=16).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=16).value = '=SUM(P%s:P%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=17).border = border
            Non_CostControl_Sheet.cell(row=row, column=17).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=17).value = '=SUM(Q%s:Q%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=18).border = border
            Non_CostControl_Sheet.cell(row=row, column=18).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=18).value = '=SUM(R%s:R%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=19).border = border
            Non_CostControl_Sheet.cell(row=row, column=19).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=19).value = '=SUM(S%s:S%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=20).border = border
            Non_CostControl_Sheet.cell(row=row, column=20).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=20).value = '=SUM(T%s:T%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=21).border = border
            Non_CostControl_Sheet.cell(row=row, column=21).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=21).value = '=SUM(U%s:U%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=22).border = border
            Non_CostControl_Sheet.cell(row=row, column=22).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=22).value = '=SUM(V%s:V%s)' %(line_start, row-1)
            
            Non_CostControl_Sheet.cell(row=row, column=23).border = border
            Non_CostControl_Sheet.cell(row=row, column=23).fill = redFill
            Non_CostControl_Sheet.cell(row=row, column=23).value = '=SUM(W%s:W%s)' %(line_start, row-1)
            
        stream1 = cStringIO.StringIO()
        workbook.save(stream1)
        filename = '%s.xlsx' % (template_file.name)
        self.env.cr.execute(""" DELETE FROM budget_xls_output """)
        attach_id = self.env['budget.xls.output'].create(
            {'name': filename,
             'xls_output': base64.encodestring(stream1.getvalue()),
             }
        )
        return attach_id

    @api.multi
    def export_budget(self):
        budget_ids = self.env.context.get('active_ids')
        attach_id = self.update_budget_xls(budget_ids, self.attachment_id.id)
        return {
            'context': self.env.context,
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'budget.xls.output',
            'res_id': attach_id.id,
            'type': 'ir.actions.act_window',
            'target': 'new',
        }
